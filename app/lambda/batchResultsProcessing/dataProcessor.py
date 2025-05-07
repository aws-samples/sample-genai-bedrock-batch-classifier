import json
import os
import logging
import io, csv
import re
from typing import Dict, List, Optional, Tuple
from typing import Dict, List, Optional
from utils.dynamodb import get_job_status_items, update_or_create_job_status_record
from utils.id_generator import get_current_date_short_str
from utils.s3 import save_file_to_s3
from batchResultsProcessing.environmentConfig import EnvironmentConfig
from openpyxl import Workbook

# Configure logging
logger = logging.getLogger(__name__)
log_level = os.environ.get("LOG_LEVEL", "INFO")
logger.setLevel(log_level)


class DataProcessor:
    """Processes classification results from the batch classifier."""

    def __init__(self, config: EnvironmentConfig):
        """
        Initialize ResultsProcessor.

        Args:
            config: Environment configuration
        """
        self.config = config

    def process_results(self, content: List[str]) -> Optional[List[Dict]]:
        """
        Process batch classification results.

        Args:
            content: List of JSONL content strings

        """
        try:
            records = []
            for line in content:
                data = json.loads(line.strip())
                record_id = data["recordId"]
                input_text = data["modelInput"]["messages"][0]["content"][0]["text"]
                output_result = data["modelOutput"]["content"][0]["text"]

                class_content, rationale_content = self._extract_class_and_rationale(output_result)

                records.append({
                    "id": record_id,
                    "input_text": input_text,
                    "class": class_content,
                    "rationale": rationale_content
                })

            logger.info(f"Processed {len(records)} classification records")
            return records
        except Exception as e:
            logger.error(f"Error processing results: {e}")
            return None

    def check_if_all_jobs_completed(self, parent_id) -> bool:
        """
        Check if all jobs are completed.

        Args:
            parent_id (str): The parent ID to check
        """
        try:
            job_status_table = self.config.get("job_status_table")

            response = get_job_status_items(
                job_status_table,
                {"parent_id": parent_id},
                page_size=25,
                consistent_read=True
            )
            if response:
                for item in response:
                    job_status = item["job_status"]["S"]
                    job_id = item["id"]["S"]
                    if job_status != "COMPLETED":
                        logger.info(f"Job {job_id} is still running and has Bedrock status: {job_status}")
                        return False
                    else:
                        logger.info(f"Job {job_id} is completed")
                return True
            else:
                return False
        except Exception as e:
            logger.error(f"Error checking if all jobs are completed: {e}")
            return False

    def save_results_externally(self, parent_job_id: str, base_filename: str, records: List[Dict]) -> None:
        """
        Save processed results to external S3.

        Args:
            parent_job_id (str): parent id that groups batches together
            base_filename (sr): item_id or the name of the output file
            records (List[Dict]): List of processed records

        """
        try:
            current_date = get_current_date_short_str()
            output_bucket_name = self.config.get("output_bucket_name")
            output_format = self.config.get("output_format")
            output_folder_name = self.config.get("output_folder_name")
            
            # Generate output paths
            output_key = f"{output_folder_name}/{current_date}/{parent_job_id}/{base_filename}{output_format}"

            # Convert and save results
            file_content = self._convert_to_json(records)
            if not file_content:
                return False

            if output_format == ".csv":
                csv_content = self._convert_to_csv(records)
                if not csv_content:
                    return False
                save_file_to_s3(csv_content, output_bucket_name, output_key)
            elif output_format == ".json":
                save_file_to_s3(file_content, output_bucket_name, output_key)
            elif output_format == ".xlsx":
                excel_content = self._convert_to_excel(records)
                if excel_content is None:
                    logger.error("Failed to convert records to Excel format")
                    return False
                
                save_file_to_s3(excel_content, output_bucket_name, output_key)

        except Exception as e:
            logger.error(f"Error saving external results: {e}")
    
    def save_results_internally(self, internal_bucket_name: str, parent_job_id: str, item_id: str, records: List[Dict]) -> None:
        """
        Save processed results to internal S3.

        Args:
            internal_bucket_name (str): the bucket name that will be used for saving the processed records
            parent_job_id (str): parent id that groups batches together
            base_filename (sr): item_id or the name of the output file
            records (List[Dict]): List of processed records

        """
        try:
            current_date = get_current_date_short_str()
            internal_processed_folder = self.config.get("internal_processed_folder")
            
            # Generate output paths
            internal_key = f"{internal_processed_folder}/{current_date}/{parent_job_id}/{item_id}.json"

            # Convert and save results
            file_content = self._convert_to_json(records)
            if not file_content:
                return False

            # Save internal copy
            save_file_to_s3(file_content, internal_bucket_name, internal_key)

        except Exception as e:
            logger.error(f"Error saving results internally: {e}")

    def update_job_status(self, parent_job_id: str, item_id: str) -> None:
        """
        Update job status in DynamoDB.

        Args:
            parent_job_id(str): Parent ID that groups batches together
            item_id (str): The DynamoDB item ID to update
        """
        try:
            job_status_table = self.config.get("job_status_table")

            update_or_create_job_status_record(
                job_status_table,
                item_id,
                {"job_status": "COMPLETED"}
            )

            if self.check_if_all_jobs_completed(parent_job_id):
                logger.info(f"All jobs for parent {parent_job_id} are completed")

        except Exception as e:
            logger.error(f"Error updating job status: {e}")

    @staticmethod
    def _extract_class_and_rationale(text: str) -> Tuple[str, str]:
        """
        Extract classification and rationale from model output.

        Args:
            text: Raw model output text

        """
        pattern = r"<class>(.*?)</class>\s*(.*)"
        match = re.search(pattern, text, re.DOTALL)
        
        if match:
            class_content = match.group(1).strip()
            rationale_content = match.group(2).strip()
        else:
            class_content = "Classification was not successful."
            rationale_content = "No rationale found."

        return class_content, rationale_content

    @staticmethod
    def _convert_to_csv(records: List[Dict]) -> Optional[str]:
        """
        Convert records to CSV format.

        Args:
            records: List of processed records

        """
        try:
            csv_buffer = io.StringIO()
            fieldnames = records[0].keys() if records else []
            writer = csv.DictWriter(csv_buffer, fieldnames=fieldnames)
            writer.writeheader()
            for record in records:
                writer.writerow(record)
            return csv_buffer.getvalue()
        except Exception as e:
            logger.error(f"Error converting to CSV: {e}")
            return None

    @staticmethod
    def _convert_to_excel(records: List[Dict]) -> Optional[io.BytesIO]:
        """
        Convert records to Excel format.
        
        Args:
            records: List of dictionaries containing the data
        Returns:
            BytesIO: Excel file content as a BytesIO object, or None if conversion fails
        """
        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            
            # Write headers if records exist
            if records:
                headers = list(records[0].keys())
                for col, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col, value=str(header))
            
                # Write data rows
                for row_idx, record in enumerate(records, start=2):
                    for col_idx, (key, value) in enumerate(record.items(), start=1):
                        # Handle None values and convert all values to string
                        cell_value = str(value) if value is not None else ""
                        ws.cell(row=row_idx, column=col_idx, value=cell_value)

            # Save to bytes buffer
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            
            # Reset buffer position
            excel_buffer.seek(0)
            return excel_buffer
            
        except Exception as e:
            logger.error(f"Error converting to Excel: {str(e)}")
            return None


    @staticmethod
    def _convert_to_json(records: List[Dict]) -> Optional[str]:
        """
        Convert records to JSONL format.

        Args:
            records: List of processed records

        """
        try:
            if not records:
                logger.info("No records to convert")
                return None

            return "\n".join(json.dumps(record, ensure_ascii=False) for record in records)
        except Exception as e:
            logger.error(f"Error converting to JSON: {e}")
            return None